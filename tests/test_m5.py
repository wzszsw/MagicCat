"""M5 测试：表导出（CSV/Excel/JSON/SQL）与 CSV 导入。"""

from __future__ import annotations

import csv
import json
import time

from magiccat.models.profile import DEFAULT_GROUP, ConnectionProfile
from magiccat.services import transfer


def _profile(mysql_env: dict) -> ConnectionProfile:
    return ConnectionProfile(name="M5", group=DEFAULT_GROUP, database="test",
                             host=mysql_env["host"], port=mysql_env["port"],
                             username=mysql_env["user"], password=mysql_env["password"])


def _setup(connection_service, profile, q) -> tuple[str, str]:
    """建源表并灌入含边界字符的数据；返回 (src, dst)。"""
    suffix = int(time.time() * 1000)
    src = f"mc_m5_src_{suffix}"
    dst = f"mc_m5_dst_{suffix}"
    q.execute(profile, (
        f"CREATE TABLE `{src}` (id INT PRIMARY KEY AUTO_INCREMENT, "
        "name VARCHAR(50) NOT NULL, note VARCHAR(100) NULL, price DECIMAL(10,2) NULL)"
        " ENGINE=InnoDB"))
    q.execute(profile, (
        f"CREATE TABLE `{dst}` LIKE `{src}`"))
    q.execute(profile, (
        f"INSERT INTO `{src}` (name, note, price) VALUES "
        "('奥力给', '带,逗号 and \"引号\"', 1.50),"
        "('slash\\\\back', NULL, NULL),"
        "('换行\n在里面', 'semi;colon', 0.99)"))
    return src, dst


def test_export_all_formats(tmp_path, mysql_env, connection_service):
    from magiccat.services.data_service import DataService
    from magiccat.services.metadata_service import MetadataService
    from magiccat.services.query_service import QueryService

    profile = _profile(mysql_env)
    connection_service.add(profile)
    q = QueryService(connection_service)
    data = DataService(connection_service)
    meta = MetadataService(connection_service)
    src, dst = _setup(connection_service, profile, q)
    try:
        # CSV
        csv_path = tmp_path / "out.csv"
        res = transfer.export_table(profile, "test", src, csv_path, "csv", data, meta)
        assert res["rows"] == 3 and not res["cancelled"]
        with open(csv_path, encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f))
        assert rows[0] == ["id", "name", "note", "price"]
        assert len(rows) == 4
        assert any(r[1] == "奥力给" and r[2].startswith("带,逗号") for r in rows[1:])
        assert any(r[2] == "" for r in rows[1:])  # NULL -> 空串

        # Excel
        xlsx_path = tmp_path / "out.xlsx"
        assert transfer.export_table(profile, "test", src, xlsx_path, "excel", data, meta)["rows"] == 3
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        ws = wb.active
        assert ws.max_row == 4 and ws.max_column == 4
        wb.close()

        # JSON
        json_path = tmp_path / "out.json"
        assert transfer.export_table(profile, "test", src, json_path, "json", data, meta)["rows"] == 3
        objs = json.loads(json_path.read_text(encoding="utf-8"))
        assert len(objs) == 3
        assert objs[0]["name"] == "奥力给" and objs[1]["note"] is None

        # SQL：含 CREATE + 3 条 INSERT，特殊字符转义后可回灌
        sql_path = tmp_path / "out.sql"
        assert transfer.export_table(profile, "test", src, sql_path, "sql", data, meta)["rows"] == 3
        text = sql_path.read_text(encoding="utf-8")
        assert "CREATE TABLE" in text and "PRIMARY KEY (`id`)" in text
        assert text.count("INSERT INTO") == 3
        # 先删表再回灌，验证 CREATE+INSERT 可完整重建
        q.execute(profile, f"DROP TABLE IF EXISTS `{src}`")
        results = q.execute(profile, text)
        assert all(r["kind"] == "update" for r in results)
        rows = q.execute(profile, f"SELECT name FROM `{src}` ORDER BY id")
        assert [r[0] for r in rows[0]["rows"]] == ["奥力给", "slash\\back", "换行\n在里面"]
    finally:
        q.execute(profile, f"DROP TABLE IF EXISTS `{src}`")
        q.execute(profile, f"DROP TABLE IF EXISTS `{dst}`")
        connection_service.close(profile.id)


def test_import_csv(tmp_path, mysql_env, connection_service):
    from magiccat.services.data_service import DataService
    from magiccat.services.metadata_service import MetadataService
    from magiccat.services.query_service import QueryService

    profile = _profile(mysql_env)
    connection_service.add(profile)
    q = QueryService(connection_service)
    meta = MetadataService(connection_service)
    src, dst = _setup(connection_service, profile, q)
    try:
        # 先把 src 导成 CSV，再导入 dst（同名列映射）
        csv_path = tmp_path / "data.csv"
        transfer.export_table(profile, "test", src, csv_path, "csv",
                              DataService(connection_service), meta)
        res = transfer.import_csv(profile, "test", dst, csv_path, q, meta,
                                  has_header=True, empty_as_null=True)
        assert res["rows"] == 3 and res["first_error"] is None

        rows = q.execute(profile, f"SELECT name, note FROM `{dst}` ORDER BY id")
        values = rows[0]["rows"]
        assert len(values) == 3
        assert values[0] == ["奥力给", "带,逗号 and \"引号\""]
        assert values[1] == ["slash\\back", None]  # 空 note -> NULL
        assert values[2][1] == "semi;colon"
    finally:
        q.execute(profile, f"DROP TABLE IF EXISTS `{src}`")
        q.execute(profile, f"DROP TABLE IF EXISTS `{dst}`")
        connection_service.close(profile.id)
